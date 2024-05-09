import os
from dotenv import load_dotenv, find_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

load_dotenv(find_dotenv())


class ExcelWriter:
    def __init__(self):
        self.file_path = os.getenv('XLSX_PATH')
        self.wb = None
        self.ws = None

    def create_new_book(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def add_header(self, header):
        self.ws.append(header)
        for cell in self.ws[1]:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    def add_data(self, data):
        for row in data:
            self.ws.append(row)

    def set_column_width(self, width):
        for col in range(1, self.ws.max_column + 1):
            col_letter = get_column_letter(col)
            self.ws.column_dimensions[col_letter].width = width

    def save(self):
        self.wb.save(self.file_path)
